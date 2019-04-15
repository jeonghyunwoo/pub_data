from bs4 import BeautifulSoup as soup
from urllib.request import urlopen
from urllib.parse import quote_plus
import pandas as pd
import numpy as np
from newspaper import Article
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import time
import os

def news(qry):
    lks = []
    qryb = quote_plus(qry)
    baseurl = "https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query="+qryb+"&sm=tab_pge&sort=0&photo=0&field=0&reporter_article=&pd=0&ds=&de=&docid=&nso=so:r,p:all,a:all&mynews=0&cluster_rank=48&start={}&refresh_start=0"
    urls = [baseurl.format(i*10+1) for i in range(0,3)]
    aa = list(map(lambda x: soup(urlopen(x),'html.parser'),urls))
    lks = [a.find_all('a',{'class':'_sp_each_title'}) for a in aa]
    #a = soup(urlopen(url),'html.parser')
    #lk = a.find_all('a',{'class':'_sp_each_title'})
    href = []
    date = []
    title = []
    text = []
    for lk in lks:
        for link in lk:
            h = link.get('href')
            href.append(h)
            try:
                nw = Article(h)
                nw.download()
                nw.parse()
                title.append(nw.title)
                text.append(nw.text)
                dat = nw.publish_date
                dat = dat.strftime('%Y%m%d')
                date.append(dat)
            except Exception as e:
                print('no date')
                date.append(np.nan)
            time.sleep(np.random.rand())

    res = pd.DataFrame(dict(title=title,
                            date=date,
                            text=text,
                            url=href))

    wb = Workbook()
    ws = wb.active
    colname = ['title','date','text','url']

    for i in range(0,4):
        ws.cell(1,i+1).value = colname[i]

    for i in range(0,len(title)):
        ws.cell(i+2,1).value = title[i]
        ws.cell(i+2,2).value = date[i]
        ws.cell(i+2,3).value = text[i]
        ws.cell(i+2,4).hyperlink = href[i]
        ws.cell(i+2,4).font = Font(underline='single', color='0563C1')

    wb.save('news.xlsx')
    os.startfile('news.xlsx')
